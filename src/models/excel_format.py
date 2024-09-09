import openpyxl as op
from openpyxl.styles import Font, PatternFill, Alignment


def excel_format(file_path: str, validator: str) -> None:
    """
    Formatea un archivo Excel aplicando estilos según el validador.
    
    Parámetros:
    file_path (str): La ruta del archivo Excel que se desea formatear.
    validator: Una función u objeto que determina las reglas de formato que se deben aplicar.
    
    Retorna:
    None
    """
    workbook = op.load_workbook(file_path)
    sheet = workbook.active
    
    # Definir estilos generales
    general_font = Font(name='Calibri', size=11)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    alignment_center = Alignment(horizontal='center', vertical='center')
    
    # Definir estilos específicos
    header_font_white = Font(name='Calibri', size=11, bold=True, color='000000')
    header_fill_blue = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    header_fill_yellow = PatternFill(start_color='FFD965', end_color='FFD965', fill_type='solid')
    header_fill_green = PatternFill(start_color='C4D79B', end_color='C4D79B', fill_type='solid')
    header_fill_orange = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')
    
    # Aplicar estilos generales a todas las celdas
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = general_font
    
    # Aplicar estilos al encabezado (fila 1)
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = alignment_center
    
    # Aplicar estilos específicos según el validador
    if validator == 'mono':
        for col in range(1, 10):  # Columnas A-I
            sheet.cell(row=1, column=col).fill = header_fill_blue
        for col in range(10, 14):  # Columnas J-M
            sheet.cell(row=1, column=col).fill = header_fill_yellow
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(14, 16):  # Columnas N-O
            sheet.cell(row=1, column=col).fill = header_fill_green
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(16, 17):  # Columna P
            sheet.cell(row=1, column=col).fill = header_fill_orange
            sheet.cell(row=1, column=col).font = header_font_white
    
    elif validator == 'multi':
        for col in range(1, 10):  # Columnas A-I
            sheet.cell(row=1, column=col).fill = header_fill_blue
        for col in range(10, 12):  # Columnas J-K
            sheet.cell(row=1, column=col).fill = header_fill_green
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(12, 15):  # Columnas L-N
            sheet.cell(row=1, column=col).fill = header_fill_yellow
            sheet.cell(row=1, column=col).font = header_font_white
    
    elif validator == 'multi_final':
        for col in range(1, 10):  # Columnas A-I
            sheet.cell(row=1, column=col).fill = header_fill_blue
        for col in range(10, 12):  # Columnas J-K
            sheet.cell(row=1, column=col).fill = header_fill_green
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(12, 16):  # Columnas L-P
            sheet.cell(row=1, column=col).fill = header_fill_yellow
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(16, 17):  # Columna P
            sheet.cell(row=1, column=col).fill = header_fill_orange
            sheet.cell(row=1, column=col).font = header_font_white
    
    elif validator == 'instruccion':
        for col in range(1, 11): # Columnas A-J
            sheet.cell(row=1, column=col).fill = header_fill_blue
        for col in range(11, 13): # Columnas K-L
            sheet.cell(row=1, column=col).fill = header_fill_green
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(13, 17):  # Columnas M-P
            sheet.cell(row=1, column=col).fill = header_fill_yellow
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(17, 18):  # Columna Q
            sheet.cell(row=1, column=col).fill = header_fill_orange
            sheet.cell(row=1, column=col).font = header_font_white
    
    workbook.save(file_path)